# zaiko-v7-2w.py
import datetime
import io
from copy import copy
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter

import streamlit as st


# ----------------------------
# Utilities
# ----------------------------
def find_sheet_by_strip(workbook, target_name: str):
    t = target_name.strip()
    for name in workbook.sheetnames:
        if name.strip() == t:
            return workbook[name]
    return None


def parse_target_date(target_date_str: str) -> datetime.date:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(target_date_str, fmt).date()
        except ValueError:
            continue
    raise ValueError("日付形式が無効です。YYYY-MM-DD または YYYY/MM/DD 形式で入力してください。")


def normalize_excel_cell_to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def resolve_honzan_col_letter_2w(ws_input, target_date: datetime.date) -> str:
    """
    2週間ver（確定仕様）
    - 在庫集計表の 5行目 で target_date と一致する列を探す
    - その列 + 8 が「本残」列（=在庫表 C列：前夜本残）
    """
    DATE_ROW = 5
    HEADER_ROW = 7

    date_col_idx = None
    for c in range(1, ws_input.max_column + 1):
        v = ws_input.cell(row=DATE_ROW, column=c).value
        v_date = normalize_excel_cell_to_date(v)
        if v_date == target_date:
            date_col_idx = c
            break

    if date_col_idx is None:
        raise ValueError(f"在庫集計表の{DATE_ROW}行目に {target_date.strftime('%Y-%m-%d')} が見つかりません。")

    honzan_col_idx = date_col_idx + 8

    # 安全装置：+8先が本残列か確認
    header_val = ws_input.cell(row=HEADER_ROW, column=honzan_col_idx).value
    header_str = "" if header_val is None else str(header_val)
    if "本残" not in header_str:
        raise ValueError(
            f"{target_date.strftime('%Y-%m-%d')} の本残列が見つかりません。"
            f"（{get_column_letter(honzan_col_idx)}{HEADER_ROW}='{header_str}'）"
        )

    return get_column_letter(honzan_col_idx)


def remove_xlm_defined_names(wb: openpyxl.Workbook):
    """
    Excel が「マクロ有効コンテンツ」と見なす定義名を除去する。
    2週間テンプレに入っている _xleta.VLOOKUP (xlm=1) が主犯。
    """
    # keys を先に固定
    for name in list(wb.defined_names):
        dn = wb.defined_names.get(name)
        # dn が list の場合もあるが、今回のテンプレは単体なので単体前提でOK。
        # 念のため list でも見る。
        try:
            if str(name).startswith("_xleta."):
                wb.defined_names.pop(name, None)
                continue

            if isinstance(dn, list):
                # どれか1つでも xlm=1 なら削除
                if any(getattr(x, "xlm", None) in (True, "1", 1) for x in dn):
                    wb.defined_names.pop(name, None)
                continue

            if getattr(dn, "xlm", None) in (True, "1", 1):
                wb.defined_names.pop(name, None)
                continue
        except Exception:
            # 変な定義名がいても落とさない
            continue


def keep_only_two_sheets(wb: openpyxl.Workbook, keep_titles: set[str]):
    keep_norm = {t.strip() for t in keep_titles}
    for ws in list(wb.worksheets):
        if ws.title.strip() not in keep_norm:
            wb.remove(ws)
    if wb.worksheets:
        wb.active = 0


# ----------------------------
# Main processing
# ----------------------------
def create_categorized_inventory_excel(uploaded_file, target_date_str: str):
    INPUT_SHEET = "在庫集計表"
    OUT_BOX = "在庫表（箱）"
    OUT_SMALL = "在庫表（こもの）"

    try:
        target_date = parse_target_date(target_date_str)
    except ValueError as e:
        return f"エラー: {e}"

    # 1) データ抽出用（data_only=True）
    try:
        uploaded_file.seek(0)
        wb_input = openpyxl.load_workbook(uploaded_file, data_only=True, keep_vba=False)
    except Exception as e:
        return f"エラー: 入力ファイルの読み込みに失敗しました: {e}"

    if INPUT_SHEET not in wb_input.sheetnames:
        return f"エラー: シート『{INPUT_SHEET}』が見つかりません。"

    ws_input = wb_input[INPUT_SHEET]

    # 本残列（ターゲット日付列 + 8）
    try:
        col_letter = resolve_honzan_col_letter_2w(ws_input, target_date)
    except ValueError as e:
        return f"エラー: {e}"

    HEADER_ROW = 7
    exclusion_keywords = [
        "配達料", "運賃", "カステラ", "十勝の息吹",
        "有機納豆", "ひきわり", "豆腐", "丸大豆"
    ]
    exclusion_toichi = "東一"

    boxed, smalls = [], []

    for r in range(HEADER_ROW + 1, ws_input.max_row + 1):
        code = ws_input.cell(row=r, column=column_index_from_string("A")).value
        name = ws_input.cell(row=r, column=column_index_from_string("B")).value

        if (code is None or str(code).strip() == "") and (name is None or str(name).strip() == ""):
            continue
        if not isinstance(name, str):
            continue

        name_lower = name.lower()
        if any(kw.lower() in name_lower for kw in exclusion_keywords):
            continue

        val = ws_input.cell(row=r, column=column_index_from_string(col_letter)).value
        if val is None or val == "":
            val = 0

        if exclusion_toichi in name and (val == 0 or val == "0"):
            continue

        rec = [code, name, val]
        if name.startswith("■"):
            boxed.append(rec)
        else:
            smalls.append(rec)

    # ▢優先 → 商品コード昇順
    def sort_key(row):
        c, n, _ = row
        n_s = "" if n is None else str(n).strip()
        c_s = "" if c is None else str(c)
        return (not n_s.startswith("▢"), c_s)

    smalls.sort(key=sort_key)

    # 保険：完全空行を落とす
    def compact(data):
        out = []
        for c, n, v in data:
            c_s = "" if c is None else str(c).strip()
            n_s = "" if n is None else str(n).strip()
            if c_s == "" and n_s == "" and (v is None or v == "" or v == 0):
                continue
            out.append([c, n, v])
        return out

    boxed = compact(boxed)
    smalls = compact(smalls)

    # 2) 書き込み用（data_only=False）
    # ★ここが重要：keep_vba=False で読み直す（= xlsm要素を持ち出さない）
    try:
        uploaded_file.seek(0)
        wb_output = openpyxl.load_workbook(uploaded_file, data_only=False, keep_vba=False)
    except Exception as e:
        return f"エラー: 出力用にテンプレートを読み込めませんでした: {e}"

    ws_box = find_sheet_by_strip(wb_output, OUT_BOX)
    ws_small = find_sheet_by_strip(wb_output, OUT_SMALL)
    if ws_box is None or ws_small is None:
        return f"エラー: 出力先シート『{OUT_BOX}』『{OUT_SMALL}』が見つかりません。"

    # 既存データクリア
    def clear_existing_data(ws):
        max_clear_row = max(ws.max_row, 2000)
        for rr in range(3, max_clear_row + 1):
            for cc in range(1, 5):  # A-D
                cell = ws.cell(rr, cc)
                cell.value = None
                if cc <= 3:
                    cell.font = Font(name="ＭＳ Ｐゴシック", size=26)

    # テンプレ行（3行目）で書式維持しながら転記
    def write(ws, data):
        clear_existing_data(ws)

        template_row = 3
        template_height = ws.row_dimensions[template_row].height
        if template_height is None:
            template_height = 39.0

        for rr, row_data in enumerate(data, start=3):
            for col_idx, value in enumerate(row_data, start=1):  # A-C
                src = ws.cell(template_row, col_idx)
                dst = ws.cell(rr, col_idx)
                dst.value = value

                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)

                # B列 shrink_to_fit 強制
                if col_idx == 2:
                    a = src.alignment
                    dst.alignment = Alignment(
                        horizontal=a.horizontal,
                        vertical=a.vertical,
                        wrap_text=a.wrap_text,
                        shrink_to_fit=True,
                    )

            ws.row_dimensions[rr].height = template_height

    def reset_print_area(ws, last_row):
        ws.print_area = f"A1:D{last_row}"

    def enforce_right_border_thin(ws, start_row, end_row, col_index):
        thin = Side(style="thin")
        tpl = ws.cell(3, col_index)
        final = Border(
            left=tpl.border.left,
            right=thin,
            top=tpl.border.top,
            bottom=tpl.border.bottom,
        )
        for rr in range(start_row, end_row + 1):
            ws.cell(rr, col_index).border = copy(final)

    def hide_trailing_rows(ws, start_row):
        last = max(ws.max_row, 2000)
        for rr in range(start_row, last + 1):
            ws.row_dimensions[rr].hidden = True

    write(ws_box, boxed)
    write(ws_small, smalls)

    last_row_box = 2 + len(boxed)
    last_row_small = 2 + len(smalls)

    if last_row_box >= 3:
        enforce_right_border_thin(ws_box, 3, last_row_box, 4)
    if last_row_small >= 3:
        enforce_right_border_thin(ws_small, 3, last_row_small, 4)

    reset_print_area(ws_box, max(last_row_box, 3))
    reset_print_area(ws_small, max(last_row_small, 3))

    hide_trailing_rows(ws_box, 3 + len(boxed) + 1)
    hide_trailing_rows(ws_small, 3 + len(smalls) + 1)

    # ★ここが主犯対策：定義名のXLMマクロ要素を削除
    remove_xlm_defined_names(wb_output)

    # ★2シート以外を落とす（必要なら）
    keep_only_two_sheets(wb_output, {OUT_BOX, OUT_SMALL})

    # 保存（xlsxで出す：ここまででマクロ扱い要素を落としている）
    out_name = f"在庫集計結果_{target_date.strftime('%Y%m%d')}.xlsx"
    out_buf = io.BytesIO()
    try:
        wb_output.save(out_buf)
    except Exception as e:
        return f"エラー: 保存に失敗しました: {e}"

    excel_data = out_buf.getvalue()
    msg = (
        "✅ 在庫集計が完了しました。\n"
        f"・箱もの：{len(boxed)}件\n"
        f"・こもの：{len(smalls)}件（▢優先ソート済み）\n"
    )
    return excel_data, out_name, msg


# ----------------------------
# Streamlit UI
# ----------------------------
st.title("📋在庫表（2週間対応版）")

uploaded_file = st.file_uploader(
    "1. 入力Excelファイル（.xlsx / .xlsm）",
    type=["xlsx", "xlsm"],
)

JST = ZoneInfo("Asia/Tokyo")
today_jp = datetime.datetime.now(JST).date()
target_date = st.date_input("2. 在庫集計日", value=today_jp)

if st.button("集計してExcel生成", key="generate_excel"):
    if uploaded_file is None:
        st.error("入力Excelファイルを選択してください。")
    else:
        date_str = target_date.strftime("%Y-%m-%d")
        with st.spinner("処理中... Excelファイルを生成しています。"):
            result = create_categorized_inventory_excel(uploaded_file, date_str)

        if isinstance(result, str):
            st.error(result)
        else:
            excel_data, file_name, msg = result
            st.success(msg)
            st.download_button(
                label="📁 集計結果をダウンロード",
                data=excel_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
