import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Border, Side
import datetime
# import tkinter as tk # --- 削除 ---
# from tkinter import filedialog, messagebox, ttk # --- 削除 ---
import os
from copy import copy
# from tkcalendar import DateEntry # --- 削除 ---

import streamlit as st  # --- 追加 ---
import io               # --- 追加 ---
from zoneinfo import ZoneInfo # --- ★修正: タイムゾーンライブラリをインポート ---


def find_sheet(workbook, target_name):
    """シート名をスペース無視で照合して取得"""
    for name in workbook.sheetnames:
        if name.strip() == target_name.strip():
            return workbook[name]
    return None


def create_categorized_inventory_excel(
    input_file_buffer,  # 変更: ファイルパス -> ファイルバッファ
    target_date_str: str
): # 変更: 戻り値の型が変わる (エラー時はstr, 成功時はTuple)
    """
    在庫集計ファイルから「箱もの」「こもの」を分類し、
    既存シート「在庫表（箱）」「在庫表（こもの）」に転記。
    書式は保持しつつ、空白行や残データを確実に除外。
    """

    input_sheet_name = "在庫集計表"  # 固定シート名

    try:
        # --- 日付解析 ---
        try:
            target_date = datetime.datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return "エラー: 日付形式が無効です。YYYY-MM-DD形式で入力してください。"

        # --- 曜日→列マッピング（日曜=M） ---
        ordered_days_jp = ['日曜日', '月曜日', '火曜日', '水曜日', '木曜日', '金曜日']
        ordered_columns_letters = ['M', 'R', 'W', 'AB', 'AG', 'AL']
        python_weekday_to_jp_day = {
            0: '月曜日', 1: '火曜日', 2: '水曜日',
            3: '木曜日', 4: '金曜日', 5: '土曜日', 6: '日曜日'
        }

        base_day_jp = python_weekday_to_jp_day[target_date.weekday()]
        if base_day_jp == "土曜日":
            return "エラー: 土曜日は集計対象外です。"

        col_letter = ordered_columns_letters[ordered_days_jp.index(base_day_jp)]

        # --- 入力ファイル読込 (データ取得用) ---
        # 変更: ファイルパスからバッファを読み込む
        wb_input = openpyxl.load_workbook(input_file_buffer, data_only=True)
        if input_sheet_name not in wb_input.sheetnames:
            return f"エラー: シート『{input_sheet_name}』が見つかりません。"
        ws_input = wb_input[input_sheet_name]

        header_row = 7
        exclusion_keywords = ["配達料", "運賃", "カステラ", "十勝の息吹", "有機納豆", "ひきわり", "豆腐", "丸大豆"]
        exclusion_toichi = "東一"

        boxed, smalls = [], []

        # --- データ抽出 ---
        for r in range(header_row + 1, ws_input.max_row + 1):
            code = ws_input.cell(row=r, column=column_index_from_string('A')).value
            name = ws_input.cell(row=r, column=column_index_from_string('B')).value

            if (code is None or str(code).strip() == "") and (name is None or str(name).strip() == ""):
                continue
            if not isinstance(name, str):
                continue

            name_lower = name.lower()
            if any(k.lower() in name_lower for k in exclusion_keywords):
                continue
            val = ws_input.cell(row=r, column=column_index_from_string(col_letter)).value or 0
            if exclusion_toichi.lower() in name_lower and val == 0:
                continue

            record = [code, name, val]
            if name.strip().startswith("■"):
                boxed.append(record)
            else:
                smalls.append(record)

        if not boxed and not smalls:
            return "有効なデータが見つかりません。"

        # --- 「こもの」ソート処理 ---
        def sort_komono_data(data):
            def sort_key(row):
                product_name = str(row[1]) if row[1] else ""
                product_code = str(row[0]) if row[0] else ""
                starts_with_square = product_name.strip().startswith("▢")
                return (0 if starts_with_square else 1, product_code)
            return sorted(data, key=sort_key)

        smalls = sort_komono_data(smalls)

        # --- 空行除去 ---
        def compact_records(records):
            cleaned = []
            for code, name, val in records:
                code_s = "" if code is None else str(code).strip()
                name_s = "" if name is None else str(name).strip()
                if (code_s == "" and name_s == "" and (val is None or val == "" or val == 0)):
                    continue
                cleaned.append([code, name, val])
            return cleaned

        boxed = compact_records(boxed)
        smalls = compact_records(smalls)

        # --- 出力ファイル名生成 ---
        output_file_name = f"在庫集計結果_{target_date.strftime('%Y%m%d')}.xlsx"

        # --- 出力先シート取得 (書式保持用) ---
        input_file_buffer.seek(0) 
        wb_output = openpyxl.load_workbook(input_file_buffer, data_only=False)
        ws_box = find_sheet(wb_output, "在庫表（箱）")
        ws_small = find_sheet(wb_output, "在庫表（こもの）")

        if ws_box is None or ws_small is None:
            existing = ", ".join(wb_output.sheetnames)
            return f"エラー: 出力先シートが見つかりません。\n現在のシート一覧: {existing}"

        # --- 既存データクリア ---
        def clear_existing_data(ws):
            max_clear_row = max(ws.max_row, 2000)
            for r in range(3, max_clear_row + 1):
                for c in range(1, 5): # 1(A)から4(D)まで
                    cell = ws.cell(r, c)
                    cell.value = None
                    if c <= 3: # A, B, C列のみフォントをリセット
                        cell.font = openpyxl.styles.Font(name='ＭＳ Ｐゴシック', size=26)

        # --- データ転記 ---
        def write(ws, data):
            clear_existing_data(ws)
            template_row = 3
            template_height = ws.row_dimensions[template_row].height
            if template_height is None:
                template_height = 39.0 # フォールバック

            for i, row_data in enumerate(data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    src_cell = ws.cell(template_row, col_idx) 
                    dest_cell = ws.cell(i, col_idx)
                    dest_cell.value = value
                    
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)
                    
                    if col_idx == 2: # B列のみ 'Shrink to fit' を強制
                        dest_cell.alignment = openpyxl.styles.Alignment(
                            horizontal=src_cell.alignment.horizontal,
                            vertical=src_cell.alignment.vertical,
                            shrink_to_fit=True
                        )
                
                ws.row_dimensions[i].height = template_height

        # --- 印刷範囲を調整 ---
        def reset_print_area(ws, last_data_row):
            ws.print_area = f"A1:D{last_data_row}"

        # --- 指定列の右罫線を強制追加 ---
        def enforce_right_border(ws, start_row, end_row, col_index):
            thin = Side(style="thin")
            template_cell = ws.cell(3, col_index)
            template_left = template_cell.border.left
            template_top = template_cell.border.top
            template_bottom = template_cell.border.bottom
            
            final_border = Border(
                left=template_left,
                right=thin,
                top=template_top,
                bottom=template_bottom
            )
            
            for r in range(start_row, end_row + 1):
                cell = ws.cell(r, col_index)
                cell.border = copy(final_border) 

        # --- 転記実行 ---
        write(ws_box, boxed)
        write(ws_small, smalls)

        # --- D列右罫線再描画 ---
        last_row_small = 2 + len(smalls)
        if last_row_small >= 3:
            enforce_right_border(ws_small, 3, last_row_small, 4)

        last_row_box = 2 + len(boxed)
        if last_row_box >= 3:
            enforce_right_border(ws_box, 3, last_row_box, 4)

        # --- 印刷範囲を再設定 ---
        reset_print_area(ws_box, 2 + len(boxed))
        reset_print_area(ws_small, 2 + len(smalls))

        # --- 余分な行を非表示 ---
        def hide_trailing_rows(ws, start_row):
            last = max(ws.max_row, 2000)
            for r in range(start_row, last + 1):
                ws.row_dimensions[r].hidden = True

        hide_trailing_rows(ws_box, 3 + len(boxed) + 1)
        hide_trailing_rows(ws_small, 3 + len(smalls) + 1)
        
        # --- 成功メッセージ構築 ---
        success_message = (
            f"✅ 在庫集計が完了しました。\n"
            f"・箱もの：{len(boxed)}件\n"
            f"・こもの：{len(smalls)}件（▢優先ソート済み）\n\n"
            f"下のボタンからダウンロードしてください。"
        )

        # --- メモリバッファに保存 ---
        output_buffer = io.BytesIO()
        wb_output.save(output_buffer)
        excel_data = output_buffer.getvalue()

        return (excel_data, output_file_name, success_message)

    except Exception as e:
        return f"予期せぬエラーが発生しました: {e}"


# --- Streamlit アプリケーション ---
st.title("📦 在庫集計")

# 1. 入力Excelファイル
uploaded_file = st.file_uploader("1. 入力Excelファイル (在庫集計表を含むファイル)", type=["xlsx", "xlsm"])

# 2. 集計基準日
# --- ★修正: 日本時間(JST)の現在日付を取得 ---
JST = ZoneInfo("Asia/Tokyo")
today_jp = datetime.datetime.now(JST).date()
target_date = st.date_input("2. 集計基準日", value=today_jp) # valueを日本時間に
# --- 修正ここまで ---

# 3. 実行ボタン
if st.button("集計してExcel生成"):
    if uploaded_file is None:
        st.error("入力Excelファイルを選択してください。")
    elif target_date is None:
        st.error("集計基準日を選択してください。")
    else:
        date_str = target_date.strftime('%Y-%m-%d')
        
        with st.spinner("処理中... Excelファイルを生成しています。"):
            result = create_categorized_inventory_excel(uploaded_file, date_str)

        # 結果のハンドリング
        if isinstance(result, str):
            st.error(result)
        else:
            excel_data, file_name, success_message = result
            st.success(success_message)
            
            st.download_button(
                label="📁 集計結果をダウンロード",
                data=excel_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )